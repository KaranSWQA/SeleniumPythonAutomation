import openpyxl

book = openpyxl.load_workbook("C:\\Users\\shett\\Desktop\\PythonDemo.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)

sheet.cell(row=2, column=2).value = "ABC"
print(sheet.cell(row=2, column=2).value)

for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "TC2":
        for j in range(2, sheet.max_column + 1):
            print(sheet.cell(row=i, column=j).value)
